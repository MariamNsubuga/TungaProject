from django.shortcuts import render,redirect
from .models import *
from .serializers import *
from rest_framework.generics import ListCreateAPIView,  RetrieveUpdateDestroyAPIView
from rest_framework.permissions import IsAuthenticated #ensure user is authenticated before they perform action
from rest_framework.authentication import *
#packages for filter,sort,order and search
# from django_filters import rest_framework as filters
# from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from .filters import *
from django.utils import timezone
#for sending email, creating csv and pdf
from django.http import FileResponse
from reportlab.pdfgen import canvas
from django.shortcuts import get_object_or_404
from django.views.generic import View
import io
from django.conf import settings
from django.http import HttpResponse
import os
#email packages
from django.core.mail import EmailMessage
from django.http import JsonResponse
from .email import *
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import status
from rest_framework.decorators import api_view
import csv
#excel packages
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone
#send email
from django.core.mail import send_mail
# from openpyxl.writer.excel import save_virtual_workbook
# Create your views here.


'''enables us to accept GET requests to list all the notes available. 
It also allows us to accept POST requests to create a new note.'''

class NoteList(ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = (SessionAuthentication, BasicAuthentication, TokenAuthentication)
    # queryset = Note.objects.all()
    queryset = Note.objects.all().order_by('-date_created') #sorts all the notes list by the latest date future dates first(-) makes it negative
    serializer_class = NoteSerializer
    # filter_backends = [DjangoFilterBackend]
    filter_fields = ('category')
    search_fields =('category')
    #ordering notes
    ordering_fields = ( 
        'date_created', 
        # 'Overdue'
        #does not make sense to order by category when we are already filtering by category
        'category'   
    )
    #filtering notes
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = NoteFilter
    #searching notes
    search_fields = ['title','body','date_created','category']
    #due dates
    # category = Note.objects.filter(date_created__lt=timezone.now(), category='overdue')
    category = Note.objects.filter(date_created__lt=datetime.now(), category='overdue')
    

'''
responds to GET requests to provide the details of a specific note as indicated by the id or primary_key.
It also responds to PUT requests to update one or more of the fields of a tnote.
It also deletes a note instance when a DELETE request is made.
'''
class NoteDetail(RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated] 
    authentication_classes = (SessionAuthentication, BasicAuthentication, TokenAuthentication)
    queryset = Note.objects.all()
    serializer_class = NoteSerializer


   

#generating pdf and csv
'''
generate PDF and CSV files of your notes.  create a PDF export view:
'''

class GeneratePDFView(APIView):
    permission_classes = [IsAuthenticated] 
    authentication_classes = (SessionAuthentication, BasicAuthentication, TokenAuthentication)
    def get(self, request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="note_list.pdf"'

        p = canvas.Canvas(response)
        notes = Note.objects.all()
        for note in notes:
            p.drawString(100, 700, f"Title: {note.title}")
            p.drawString(100, 680, f"Content: {note.body}")
            p.drawString(100, 660, f"date_created: {note.date_created}")
            p.drawString(100, 640, f"category: {note.category}")
            p.showPage()
        p.save()

        return response

class GenerateCSVView(APIView):
    permission_classes = [IsAuthenticated] 
    authentication_classes = (SessionAuthentication, BasicAuthentication, TokenAuthentication)
    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="note_list.csv"'

        writer = csv.writer(response)
        writer.writerow(['Title', 'Content','Date','Category'])
        notes = Note.objects.all()
        for note in notes:
            writer.writerow([note.title, note.body, note.date_created, note.category])

        return response


#generating excel



class GenerateExcelView(APIView):
    permission_classes = [IsAuthenticated] 
    authentication_classes = (SessionAuthentication, BasicAuthentication, TokenAuthentication)
    def get(self, request):
        # Create an Excel workbook and add a worksheet
        wb = Workbook()
        ws = wb.active
        # ws.title = "Note list"

        # Define column headers and styles
        ws.append(["Title", "Content", "Created At","Category"])
        # date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD HH:MM:SS')
        # ws.cell(row=1, column=3).style = date_style

        # Retrieve and add notes to the worksheet
        notes = Note.objects.all()
        serializer = NoteSerializer(notes, many=True)
        for item in serializer.data:
            ws.append([item['title'],item['body'], item['date_created'],item['category']])
        

        # Create an HttpResponse with the Excel content
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response[
            "Content-Disposition"
        ] = 'attachment; filename="note_list.xlsx"'

        # Save the workbook to the response
        # xlsx_data = save_virtual_workbook(wb)
        xlsx_data = wb.save(response)
        response.write(xlsx_data)

        return response


'''
share a particular note via email
'''

permission_classes = [IsAuthenticated] 
authentication_classes = (SessionAuthentication, BasicAuthentication, TokenAuthentication)
@api_view(['POST'])
def send_notes_via_email(request):
   
    # Validate the email and pk
    share_note_serializer = ShareNoteSerializer(data=request.data)
    if share_note_serializer.is_valid():
        email = share_note_serializer.validated_data['email']
        pk = share_note_serializer.validated_data['pk']
    else:
        return Response({'error': 'Invalid email or pk'}, status=400)

    try:
        # Retrieve the note associated with the provided pk
        note = Note.objects.get(id=pk)
    except Note.DoesNotExist:
        return Response({'error': 'Note not found'}, status=404)

    subject = f'Shared Note: {note.title}'
    message = f'Title: {note.title}\nContent: {note.body}\nDate: {note.date_created}\nCategory: {note.category}'

    from_email = 'mariamtestn@gmail.com'  
    recipient_list = [email]

    send_mail(subject, message, from_email, recipient_list, fail_silently=False)

    return Response({'message': 'Note sent via email'})